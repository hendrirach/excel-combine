import openpyxl, os, glob, numpy

#file utama
file_name = input("Masukkan nama file utama yang akan dijadikan file hasil penggabungan: ")
file_fullname = file_name + ".xlsx"
file_combination = openpyxl.load_workbook(file_fullname)
file_sheet = file_combination['Sheet1']

jumlah_file = len([name for name in os.listdir('.') if os.path.isfile(name)])
skip_blank = []

#parameter terkait dengan bentuk file excel
question_help = input("Apakah anda ingin dipandu untuk menyesuaikan bentuk file excel agar dapat digabungkan? y/n: ")
if question_help == "y" or question_help == "Y":
    #folder berisi file excel yang ingin digabungkan
    workbook_folder = input("Masukkan nama folder yang ingin di compile file-nya: ")
    skip_row = int(input("Berapa baris yang akan di-skip dari baris paling atas, (ex: jika melewati 3 baris maka ketik angka 3): "))
    skip_footer = int(input("Berapa baris yang akan di-skip dari baris paling bawah, (ex: jika naik 2 baris maka ketik angka 2): "))
    skip_left = int(input("Berapa kolom yang akan di-skip dari kolom paling kiri, (ex: jika mulai dari kolom C maka ketik angka 3): "))
    skip_right = int(input("Berapa kolom yang akan di-skip dari kolom paling kanan, (ex: jika mundur 1 kolom dari kolom terakhir maka ketik angka 1): "))
    skip_blank = [int(x) for x in input("Kolom yang akan dihapus karena merge atau kosong, " +
        "(ex: jika kolom merge ada di kolom D dan K maka ketik angka 4 11): ").split()]

def file_combine(folder, skip_row, skip_footer, skip_left, skip_right, skip_blank, file_sheet, x=0):
    os.chdir(workbook_folder)
    workbook_files = glob.glob("*.xlsx")
    for file in workbook_files:
        #print(x)
        if x < len(workbook_files):
            #print(file)
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook['Sheet1']
            worksheet_row = worksheet.max_row - (skip_footer - 1)
            print(worksheet_row)
            worksheet_column = worksheet.max_column - skip_right
            print(worksheet_column)
            file_row = file_sheet.max_row

            for r in range(skip_row + 1, worksheet_row, 1):
                print(r)    
                data = [worksheet.cell(row = r, column = col).value for col in range(skip_left, worksheet_column + 1, 1)]
                #hapus column kosong/merge
                data = numpy.delete(data, skip_blank)
            
                for index, value in enumerate(data):
                    file_sheet.cell(row = file_row + 1, column = index + 1).value = value
                file_row += 1
            x += 1

file_combine(workbook_folder, skip_row, skip_footer, skip_left, skip_right, skip_blank, file_sheet, x=0)        
file_combination.save(file_fullname)
